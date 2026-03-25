import os
import shutil
import re
import time
import platform
from datetime import datetime
import win32com.client as win32
import openpyxl

# ==========================================
# 1. CONFIG: 모든 설정값을 한곳에서 관리
# ==========================================
CONFIG = {
    "EXCEL_FILE": "Award_Data.xlsx",
    "SHEET_NAME": "Sheet1",
    "TEMPLATE_FILE": "Award_Template.hwp",
    "OUTPUT_DIR": "결과물",
    
    "HANGEUL_VISIBLE": True,   
    "MAX_PAGE_LIMIT": 1,       
    "WAIT_SEC": 0.1,           
    
    "FILE_NAME_FORMAT": "{문서번호}_{부서}_{이름}_표창장.hwp",

    "FIELD_MAPPING": {
        "문서번호": "doc_no",
        "표창명": "award_name",
        "부서": "department",
        "이름": "name",
        "내용": "content",
        "날짜": "award_date",
        "대표이사명": "ceo"
    },
    
    "STATUS_COL": "결과",
    "MSG_SUCCESS": "성공",
    "MSG_PARTIAL": "성공(일부누락)", 
}

# ==========================================
# 2. UTILITY: 보조 함수 모음
# ==========================================
def setup_environment():
    print(f"▶ 시스템 정보: Python {platform.python_version()} 환경")
    print("====================================================")
    print(f"🚨 [확인] '{CONFIG['EXCEL_FILE']}'이 열려있다면 반드시 닫아주세요.")
    print("====================================================")

    if not os.path.exists(CONFIG["OUTPUT_DIR"]):
        os.makedirs(CONFIG["OUTPUT_DIR"])
        print(f"📁 '{CONFIG['OUTPUT_DIR']}' 폴더를 생성했습니다.")

def format_date(raw_date):
    if isinstance(raw_date, datetime):
        return raw_date.strftime("%Y年%m月%d日")
    elif isinstance(raw_date, str) and raw_date.strip():
        try:
            dt = datetime.strptime(raw_date.strip(), "%Y-%m-%d")
            return dt.strftime("%Y年%m月%d日")
        except ValueError:
            return raw_date
    return None

def make_filename_safe(text):
    if not text: return "N/A"
    return re.sub(r'[\/:*?"<>|]', '_', str(text))

# ==========================================
# 3. MAIN LOGIC: 핵심 자동화 프로세스
# ==========================================
def main():
    setup_environment()

    current_dir = os.getcwd()
    excel_path = os.path.join(current_dir, CONFIG["EXCEL_FILE"])
    template_path = os.path.join(current_dir, CONFIG["TEMPLATE_FILE"])
    output_dir_path = os.path.join(current_dir, CONFIG["OUTPUT_DIR"])

    if not os.path.exists(excel_path) or not os.path.exists(template_path):
        print("❌ 오류: 엑셀 또는 템플릿 파일을 찾을 수 없습니다.")
        return

    try:
        wb = openpyxl.load_workbook(excel_path)
        ws = wb[CONFIG["SHEET_NAME"]]
    except Exception as e:
        print(f"❌ 엑셀 로드 에러: {e}")
        return

    # 1. 엑셀의 기존 헤더 읽기
    header = {cell.value: idx + 1 for idx, cell in enumerate(ws[1]) if cell.value}
    
    # 2. 필수 데이터 컬럼이 있는지 확인 (결과 컬럼은 제외하고 확인)
    for col in CONFIG["FIELD_MAPPING"].keys():
        if col not in header:
            print(f"❌ 오류: 엑셀 첫 줄에 '{col}' 컬럼이 없습니다.")
            return

    # 3. [핵심 업그레이드] 상태 기록 컬럼("결과") 자동 생성 로직
    status_col_name = CONFIG["STATUS_COL"]
    if status_col_name not in header:
        new_col_idx = ws.max_column + 1 # 현재 사용 중인 마지막 칸의 바로 다음 칸 번호
        ws.cell(row=1, column=new_col_idx).value = status_col_name # 첫 줄에 '결과'라고 쓰기
        header[status_col_name] = new_col_idx # 프로그램 기억장치(header)에도 추가
        print(f"💡 엑셀에 '{status_col_name}' 컬럼이 없어서 자동으로 생성했습니다.")

    print("⏳ 한글 프로그램을 실행 중입니다...")
    hwp = win32.gencache.EnsureDispatch("HWPFrame.HwpObject")
    hwp.XHwpWindows.Item(0).Visible = CONFIG["HANGEUL_VISIBLE"]

    counts = {"success": 0, "partial": 0, "fail": 0, "skip": 0}

    for row_idx in range(2, ws.max_row + 1):
        status_col_idx = header[CONFIG["STATUS_COL"]]
        status = str(ws.cell(row=row_idx, column=status_col_idx).value or "").strip()
        
        if status in [CONFIG["MSG_SUCCESS"], CONFIG["MSG_PARTIAL"]]:
            counts["skip"] += 1
            continue

        row_data = {}
        is_partial = False 
        
        for excel_col in CONFIG["FIELD_MAPPING"].keys():
            raw_val = ws.cell(row=row_idx, column=header[excel_col]).value
            if excel_col == "날짜":
                clean_val = format_date(raw_val)
            else:
                clean_val = str(raw_val or "").strip()
                
            row_data[excel_col] = clean_val
            
            if not clean_val and excel_col not in ["문서번호", "부서", "이름", "표창명"]:
                is_partial = True

        if not row_data["문서번호"]:
            print(f"\n🛑 [알림] {row_idx}행에서 빈 문서번호를 발견했습니다. 작업을 종료합니다.")
            break 

        # --- 새로 들어갈 코드 (완전 자동화) ---
        safe_row_data = {key: make_filename_safe(val) for key, val in row_data.items()}
        file_name = CONFIG["FILE_NAME_FORMAT"].format(**safe_row_data)
        # -------------------------------------

        save_path = os.path.join(output_dir_path, file_name)
        save_path = os.path.join(output_dir_path, file_name)

        print(f"▶ 처리 중: {file_name}", end=" ... ")

        try:
            shutil.copy(template_path, save_path)
            hwp.Open(save_path, "HWP", "forceopen:true")
            
            for excel_col, hwp_field in CONFIG["FIELD_MAPPING"].items():
                val = row_data[excel_col]
                if val:  
                    hwp.PutFieldText(hwp_field, val)

            time.sleep(CONFIG["WAIT_SEC"]) 

            if hwp.PageCount > CONFIG["MAX_PAGE_LIMIT"]:
                hwp.Clear(1)
                os.remove(save_path)
                raise Exception(f"{CONFIG['MAX_PAGE_LIMIT']}페이지 초과")

            hwp.Save()
            hwp.Clear(1)
            
            final_status = CONFIG["MSG_PARTIAL"] if is_partial else CONFIG["MSG_SUCCESS"]
            ws.cell(row=row_idx, column=status_col_idx).value = final_status
            
            if is_partial: counts["partial"] += 1
            else: counts["success"] += 1
            
            print(f"✅ {final_status}")

        except Exception as e:
            ws.cell(row=row_idx, column=status_col_idx).value = f"실패({str(e)[:15]})"
            counts["fail"] += 1
            print(f"❌ ({e})")

    hwp.Quit()
    try:
        wb.save(excel_path)
        print(f"\n🎉 완료! (완벽성공:{counts['success']}, 일부누락성공:{counts['partial']}, 실패:{counts['fail']}, 스킵:{counts['skip']})")
    except:
        print("\n❌ 엑셀 저장 실패 (파일을 닫고 다시 실행하세요)")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"\n🚨 치명적 오류: {e}")
    finally:
        input("\n엔터(Enter)를 누르면 창이 닫힙니다...")