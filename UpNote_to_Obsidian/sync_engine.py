"""
UpNote ↔ Obsidian 동기화 엔진
================================
실행 방법:
  python sync_engine.py              # 양방향 동기화
  python sync_engine.py --up-to-obs  # UpNote → Obsidian 단방향
  python sync_engine.py --obs-to-up  # Obsidian → UpNote 단방향
  python sync_engine.py --dry-run    # 실제 파일 변경 없이 결과만 미리보기
"""

import os
import re
import json
import shutil
import logging
import argparse
import subprocess
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _EXCEL_AVAILABLE = True
except ImportError:
    _EXCEL_AVAILABLE = False

# ─────────────────────────────────────────────
# 0. 설정 로드
# ─────────────────────────────────────────────
from sync_config import CONFIG


# ─────────────────────────────────────────────
# 1. 로깅 설정
# ─────────────────────────────────────────────
def setup_logger(log_path: Path) -> logging.Logger:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    logger = logging.getLogger("sync")
    logger.setLevel(logging.DEBUG)

    # 콘솔 핸들러 (INFO 이상)
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)
    ch.setFormatter(logging.Formatter("%(message)s"))

    # 파일 핸들러 (DEBUG 이상 전체 기록)
    fh = logging.FileHandler(log_path, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(
        logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", "%Y-%m-%d %H:%M:%S")
    )

    logger.addHandler(ch)
    logger.addHandler(fh)
    return logger


log = setup_logger(Path(CONFIG["LOG_PATH"]))


# ─────────────────────────────────────────────
# 2. 엑셀 로그 (sync_log.xlsx)
# ─────────────────────────────────────────────
# 컬럼 구성:
#   A: 날짜시간  B: 실행모드  C: 상태  D: UUID(앞 8자)  E: 노트 제목  F: 오류 메시지

_EXCEL_HEADERS = ["날짜시간", "실행모드", "상태", "UUID", "노트 제목", "오류 메시지"]

# 상태별 행 색상 (openpyxl PatternFill용 hex, 연한 톤)
_STATUS_FILL = {
    "UP→OBS":   "D6EAF8",   # 연파랑
    "OBS→UP":   "D5F5E3",   # 연초록
    "충돌":      "FEF9E7",   # 연노랑
    "스킵":      "F2F3F4",   # 연회색
    "오류":      "FDEDEC",   # 연빨강
    "DRY-RUN":  "EAF0FB",   # 연라벤더
}


class ExcelLogger:
    """
    동기화 이벤트를 sync_log.xlsx 에 행 단위로 누적 기록.
    파일이 없으면 헤더 포함 새로 생성, 있으면 기존 파일에 append.
    openpyxl 미설치 시 경고만 출력하고 조용히 비활성화.
    """

    def __init__(self, xlsx_path: Path):
        self.path = xlsx_path
        self.enabled = _EXCEL_AVAILABLE
        self._rows: list[tuple] = []   # 실행 중 누적, 마지막에 일괄 저장

        if not self.enabled:
            log.warning("⚠️  openpyxl 미설치 — 엑셀 로그 비활성화 (pip install openpyxl)")

    # ── 행 누적 ──────────────────────────────
    def record(
        self,
        mode: str,
        status: str,
        uuid: str = "",
        title: str = "",
        error: str = "",
    ):
        """동기화 이벤트 1건을 내부 버퍼에 추가"""
        self._rows.append((
            datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            mode,
            status,
            uuid[:8] if uuid else "",
            title,
            error,
        ))

    # ── 일괄 저장 ─────────────────────────────
    def flush(self):
        """버퍼에 쌓인 행을 xlsx 파일에 저장"""
        if not self.enabled or not self._rows:
            return

        try:
            self.path.parent.mkdir(parents=True, exist_ok=True)

            # 기존 파일 열기 or 새로 생성
            if self.path.exists():
                wb = openpyxl.load_workbook(self.path)
                ws = wb.active
            else:
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "동기화 로그"
                self._write_header(ws)

            # 행 추가
            for row in self._rows:
                ws.append(row)
                self._apply_row_style(ws, ws.max_row, row[2])  # row[2] = 상태

            # 열 너비 자동 조정 (최초 생성 시에만 의미있지만 매번 해도 무해)
            self._auto_width(ws)

            wb.save(self.path)
            log.debug(f"  엑셀 로그 저장: {len(self._rows)}행 → {self.path}")

        except Exception as e:
            log.warning(f"  ⚠️  엑셀 로그 저장 실패: {e}")
        finally:
            self._rows.clear()

    # ── 내부 헬퍼 ────────────────────────────
    def _write_header(self, ws):
        ws.append(_EXCEL_HEADERS)
        header_fill = PatternFill("solid", fgColor="2E4057")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for col_idx, _ in enumerate(_EXCEL_HEADERS, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"          # 헤더 고정

    def _apply_row_style(self, ws, row_num: int, status: str):
        hex_color = _STATUS_FILL.get(status, "FFFFFF")
        fill = PatternFill("solid", fgColor=hex_color)
        small_font = Font(size=10)
        for col_idx in range(1, len(_EXCEL_HEADERS) + 1):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.fill = fill
            cell.font = small_font
            cell.alignment = Alignment(vertical="center")

    @staticmethod
    def _auto_width(ws):
        col_widths = [22, 14, 10, 12, 40, 50]   # 컬럼별 고정 너비
        for col_idx, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(col_idx)].width = width


# ─────────────────────────────────────────────
# 3. UUID ↔ 제목 매핑 DB (sync_map.json)
# ─────────────────────────────────────────────
class SyncMap:
    """
    sync_map.json 구조:
    {
      "<uuid>": {
        "obs_path": "노트북명/제목.md",   # Obsidian vault 내 상대 경로
        "title": "노트 제목",
        "up_mtime": 1712345678.0,         # UpNote md 최종 수정 시각 (unix)
        "obs_mtime": 1712345999.0,        # Obsidian md 최종 수정 시각 (unix)
        "last_synced": "2024-04-06T12:00:00"
      },
      ...
    }
    """

    def __init__(self, map_path: Path):
        self.path = map_path
        self.data: dict = {}
        self._load()

    def _load(self):
        if self.path.exists():
            try:
                with open(self.path, encoding="utf-8") as f:
                    self.data = json.load(f)
            except json.JSONDecodeError:
                log.warning("⚠️  sync_map.json 파싱 실패 — 빈 맵으로 초기화합니다.")
                self.data = {}

    def save(self):
        self.path.parent.mkdir(parents=True, exist_ok=True)
        with open(self.path, "w", encoding="utf-8") as f:
            json.dump(self.data, f, ensure_ascii=False, indent=2)

    def get(self, uuid: str) -> dict | None:
        return self.data.get(uuid)

    def set(self, uuid: str, **kwargs):
        if uuid not in self.data:
            self.data[uuid] = {}
        self.data[uuid].update(kwargs)
        self.data[uuid]["last_synced"] = datetime.now().isoformat(timespec="seconds")

    def all_uuids(self) -> list[str]:
        return list(self.data.keys())


# ─────────────────────────────────────────────
# 3. 유틸리티
# ─────────────────────────────────────────────
_INVALID_CHARS = re.compile(r'[\\/*?:"<>|]')


def clean_filename(name: str) -> str:
    """파일명에 사용할 수 없는 문자 제거"""
    cleaned = _INVALID_CHARS.sub("", name).strip()
    # 길이 제한 (Windows 최대 경로 고려)
    return cleaned[:180] if cleaned else "제목없음"


def get_mtime(path: Path) -> float:
    """파일 수정 시각 (unix timestamp). 파일 없으면 0."""
    return path.stat().st_mtime if path.exists() else 0.0


def resolve_lnk(lnk_path: Path) -> Path | None:
    """Windows .lnk 파일의 대상 경로 반환"""
    try:
        result = subprocess.run(
            [
                "powershell",
                "-NoProfile",
                "-command",
                f'[Console]::OutputEncoding = [System.Text.Encoding]::UTF8; (New-Object -ComObject WScript.Shell).CreateShortcut("{lnk_path}").TargetPath',
            ],
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
            timeout=5,
        )
        target = result.stdout.strip()
        return Path(target) if target else None
    except Exception as e:
        log.debug(f"  lnk 해석 실패: {lnk_path} — {e}")
        return None


def extract_title(md_path: Path) -> str | None:
    """md 파일 상단 10줄에서 헤딩 또는 YAML title 추출"""
    try:
        with open(md_path, encoding="utf-8") as f:
            for line in (f.readline() for _ in range(10)):
                line = line.strip()
                m = re.match(r"^#{1,6}\s+(.*)", line)
                if m:
                    title = re.sub(r"#\w+\s*", "", m.group(1)).strip()
                    if title:
                        return title
                if line.lower().startswith("title:"):
                    return line.split(":", 1)[1].strip()
    except Exception as e:
        log.debug(f"  제목 추출 실패: {md_path} — {e}")
    return None


def strip_title_and_images(md_path: Path) -> str:
    """
    Obsidian 복사본 제작용:
    - 첫 번째 헤딩 줄 제거
    - 이미지 링크 제거 (텍스트만 유지)
    - 빈 줄 / <br> 전체 제거 (본문 포함)
    """
    try:
        with open(md_path, encoding="utf-8") as f:
            lines = f.readlines()
    except Exception as e:
        log.error(f"  파일 읽기 실패: {md_path} — {e}")
        return ""

    result = []
    title_removed = False
    for line in lines:
        if not title_removed and re.match(r"^#{1,6}\s+", line.strip()):
            title_removed = True
            continue

        # 이미지 링크 제거
        line = re.sub(r"!\[.*?\]\(.*?\)", "", line)

        # 빈 줄 / <br> 전체 제거 (원본 fix_titles.py 동작 유지)
        if line.strip() in ("", "<br>"):
            continue

        result.append(line)

    return "".join(result)


def read_obs_content(obs_path: Path) -> str:
    """Obsidian 파일 내용 읽기 (frontmatter upnote_uuid 포함)"""
    try:
        with open(obs_path, encoding="utf-8") as f:
            return f.read()
    except Exception as e:
        log.error(f"  Obsidian 파일 읽기 실패: {obs_path} — {e}")
        return ""


def write_obsidian_file(obs_path: Path, content: str, uuid: str, dry_run: bool):
    """
    Obsidian 파일 저장.
    YAML frontmatter에 upnote_uuid를 삽입하여 역방향 추적 가능하게 함.
    """
    frontmatter = f"---\nupnote_uuid: {uuid}\n---\n\n"

    # 기존 frontmatter가 있으면 upnote_uuid만 교체
    if content.startswith("---"):
        end = content.find("---", 3)
        if end != -1:
            existing_fm = content[3:end]
            # 이미 upnote_uuid가 있으면 교체, 없으면 추가
            if "upnote_uuid:" in existing_fm:
                existing_fm = re.sub(
                    r"upnote_uuid:.*", f"upnote_uuid: {uuid}", existing_fm
                )
            else:
                existing_fm = f"\nupnote_uuid: {uuid}" + existing_fm
            content = "---" + existing_fm + "---" + content[end + 3 :]
            frontmatter = ""  # 이미 처리됨

    final_content = frontmatter + content if frontmatter else content

    if dry_run:
        log.info(f"  [DRY-RUN] 저장 예정: {obs_path}")
        return

    obs_path.parent.mkdir(parents=True, exist_ok=True)
    with open(obs_path, "w", encoding="utf-8") as f:
        f.write(final_content)


def write_upnote_md(up_md_path: Path, obs_content: str, dry_run: bool):
    """
    Obsidian → UpNote 역방향: UpNote UUID.md 덮어쓰기.
    frontmatter(upnote_uuid 등)는 제거하고 순수 내용만 씀.
    """
    # frontmatter 제거
    content = obs_content
    if content.startswith("---"):
        end = content.find("---", 3)
        if end != -1:
            content = content[end + 3 :].lstrip("\n")

    if dry_run:
        log.info(f"  [DRY-RUN] UpNote 덮어쓰기 예정: {up_md_path}")
        return

    with open(up_md_path, "w", encoding="utf-8") as f:
        f.write(content)


def unique_obs_path(base_path: Path) -> Path:
    """동일 경로 파일이 있으면 (1), (2) 붙여 중복 방지"""
    if not base_path.exists():
        return base_path
    stem, suffix = base_path.stem, base_path.suffix
    idx = 1
    while True:
        candidate = base_path.with_name(f"{stem}({idx}){suffix}")
        if not candidate.exists():
            return candidate
        idx += 1


# ─────────────────────────────────────────────
# 4. UpNote 스캔
# ─────────────────────────────────────────────
def scan_upnote(config: dict) -> dict[str, dict]:
    """
    notebooks/ 폴더를 순회하여 {uuid: {md_path, rel_folder, title, mtime}} 반환.
    .lnk → Notes/UUID.md 해석 포함.
    """
    notebooks_dir = Path(config["UPNOTE_NOTEBOOKS_DIR"])
    notes_dir = Path(config["UPNOTE_NOTES_DIR"])
    result = {}

    if not notebooks_dir.exists():
        log.error(f"❌ UpNote notebooks 폴더를 찾을 수 없습니다: {notebooks_dir}")
        return result

    for dirpath, _, filenames in os.walk(notebooks_dir):
        lnk_files = [f for f in filenames if f.endswith(".md.lnk")]
        if not lnk_files:
            continue

        rel_folder = os.path.relpath(dirpath, notebooks_dir)

        for lnk_file in lnk_files:
            uuid = lnk_file.replace(".md.lnk", "")
            lnk_path = Path(dirpath) / lnk_file

            # 1순위: Notes/UUID.md
            md_path = notes_dir / f"{uuid}.md"
            if not md_path.exists():
                # 2순위: lnk 대상 경로
                target = resolve_lnk(lnk_path)
                if target and target.exists():
                    md_path = target
                else:
                    log.warning(f"  ⚠️  원본 없음: {lnk_file}")
                    continue

            title = extract_title(md_path) or uuid
            result[uuid] = {
                "md_path": md_path,
                "rel_folder": rel_folder,
                "title": title,
                "mtime": get_mtime(md_path),
            }

    log.debug(f"UpNote 스캔 완료: {len(result)}개 노트 발견")
    return result


# ─────────────────────────────────────────────
# 5. Obsidian 스캔
# ─────────────────────────────────────────────
def scan_obsidian(config: dict) -> dict[str, dict]:
    """
    Obsidian vault를 순회하여 upnote_uuid frontmatter가 있는 파일만 수집.
    {uuid: {obs_path, mtime}} 반환.
    UpNote → Obsidian 단방향 운용 기준이므로 신규 파일은 스캔하지 않음.
    """
    vault_dir = Path(config["OBSIDIAN_VAULT_DIR"])
    result = {}

    if not vault_dir.exists():
        log.warning(f"⚠️  Obsidian vault 폴더가 없습니다: {vault_dir}")
        return result

    for md_file in vault_dir.rglob("*.md"):
        try:
            with open(md_file, encoding="utf-8") as f:
                content = f.read(500)
        except Exception:
            continue

        if content.startswith("---"):
            end = content.find("---", 3)
            if end != -1:
                fm = content[3:end]
                m = re.search(r"upnote_uuid:\s*(\S+)", fm)
                if m:
                    uuid = m.group(1)
                    result[uuid] = {
                        "obs_path": md_file,
                        "mtime": get_mtime(md_file),
                    }

    log.debug(f"Obsidian 스캔 완료: {len(result)}개 노트 발견 (UUID 추적)")
    return result


# ─────────────────────────────────────────────
# 6. 동기화 방향 결정 로직
# ─────────────────────────────────────────────
def decide_direction(
    uuid: str,
    up_info: dict | None,
    obs_info: dict | None,
    sync_map: SyncMap,
) -> str:
    """
    반환값:
      "up_to_obs"  — UpNote가 최신 → Obsidian에 씀
      "obs_to_up"  — Obsidian이 최신 → UpNote에 씀
      "skip"       — 변경 없음
      "new_up"     — UpNote에만 있는 새 노트
      "new_obs"    — Obsidian에만 있는 새 노트 (역방향)
    """
    record = sync_map.get(uuid)

    up_mtime = up_info["mtime"] if up_info else 0.0
    obs_mtime = obs_info["mtime"] if obs_info else 0.0

    # 처음 보는 노트
    if record is None:
        if up_info and not obs_info:
            return "new_up"
        if obs_info and not up_info:
            return "new_obs"
        if up_info and obs_info:
            # 둘 다 있는데 맵에 없음 → 더 최근 쪽을 기준으로
            return "up_to_obs" if up_mtime >= obs_mtime else "obs_to_up"

    # 이전 동기화 기록이 있는 경우
    prev_up_mtime = record.get("up_mtime", 0.0)
    prev_obs_mtime = record.get("obs_mtime", 0.0)

    up_changed = up_mtime > prev_up_mtime + CONFIG["MTIME_TOLERANCE_SEC"]
    obs_changed = obs_mtime > prev_obs_mtime + CONFIG["MTIME_TOLERANCE_SEC"]

    if up_changed and obs_changed:
        # 충돌 → 더 최근 쪽 우선
        log.warning(f"  ⚡ 충돌 감지 [{uuid[:8]}]: 더 최근 수정 파일 우선 적용")
        return "up_to_obs" if up_mtime >= obs_mtime else "obs_to_up"

    if up_changed:
        return "up_to_obs"
    if obs_changed:
        return "obs_to_up"

    return "skip"


# ─────────────────────────────────────────────
# 7. UpNote → Obsidian
# ─────────────────────────────────────────────
def sync_up_to_obs(
    uuid: str,
    up_info: dict,
    sync_map: SyncMap,
    config: dict,
    dry_run: bool,
) -> bool:
    vault_dir = Path(config["OBSIDIAN_VAULT_DIR"])
    title = up_info["title"]
    rel_folder = up_info["rel_folder"]
    safe_title = clean_filename(title)

    # 기존 매핑이 있으면 경로 재사용, 없으면 새로 생성
    record = sync_map.get(uuid)
    if record and record.get("obs_path"):
        obs_path = vault_dir / record["obs_path"]
    else:
        obs_dir = vault_dir / rel_folder
        obs_path = unique_obs_path(obs_dir / f"{safe_title}.md")

    content = strip_title_and_images(up_info["md_path"])
    if not content.strip():
        log.warning(f"  ⚠️  빈 파일 건너뜀: {uuid[:8]} ({title})")
        return False

    write_obsidian_file(obs_path, content, uuid, dry_run)

    rel_obs = os.path.relpath(obs_path, vault_dir)
    sync_map.set(
        uuid,
        obs_path=rel_obs,
        title=title,
        up_mtime=up_info["mtime"],
        obs_mtime=get_mtime(obs_path),
    )

    log.info(f"  ✅ UP→OBS  [{uuid[:8]}] {rel_folder}/{safe_title}.md")
    return True


# ─────────────────────────────────────────────
# 8. Obsidian → UpNote
# ─────────────────────────────────────────────
def sync_obs_to_up(
    uuid: str,
    obs_info: dict,
    up_info: dict | None,
    sync_map: SyncMap,
    config: dict,
    dry_run: bool,
) -> bool:
    import uuid as uuid_lib

    obs_path: Path = obs_info["obs_path"]
    is_new = obs_info.get("is_new", False)
    content = read_obs_content(obs_path)

    if not content.strip():
        log.warning(f"  ⚠️  Obsidian 빈 파일 건너뜀: {obs_path.name}")
        return False

    # ── 신규 파일: UUID 발급 + frontmatter 삽입 후 처리 ──────────
    if is_new:
        new_uuid = str(uuid_lib.uuid4())
        log.info(f"  🆕 신규 Obsidian 파일 감지: {obs_path.name} → UUID 발급 [{new_uuid[:8]}]")

        # Obsidian 파일에 upnote_uuid frontmatter 삽입
        write_obsidian_file(obs_path, content, new_uuid, dry_run)

        # UpNote Notes/ 폴더에 UUID.md 생성
        notes_dir = Path(config["UPNOTE_NOTES_DIR"])
        up_md_path = notes_dir / f"{new_uuid}.md"
        write_upnote_md(up_md_path, content, dry_run)

        log.info(f"  ✅ OBS→UP (신규) [{new_uuid[:8]}] {obs_path.stem} → {up_md_path.name}")
        sync_map.set(
            new_uuid,
            obs_path=str(obs_path.relative_to(config["OBSIDIAN_VAULT_DIR"])),
            title=extract_title(obs_path) or obs_path.stem,
            up_mtime=get_mtime(up_md_path),
            obs_mtime=get_mtime(obs_path),
        )
        return True

    # ── 기존 파일: 일반 업데이트 ─────────────────────────────────
    if up_info:
        up_md_path = up_info["md_path"]
        write_upnote_md(up_md_path, content, dry_run)
        log.info(f"  ✅ OBS→UP  [{uuid[:8]}] → {up_md_path.name}")
        sync_map.set(
            uuid,
            up_mtime=get_mtime(up_md_path),
            obs_mtime=obs_info["mtime"],
        )
    else:
        notes_dir = Path(config["UPNOTE_NOTES_DIR"])
        up_md_path = notes_dir / f"{uuid}.md"
        write_upnote_md(up_md_path, content, dry_run)
        log.info(f"  ✅ OBS→UP (신규) [{uuid[:8]}] → {up_md_path.name}")
        sync_map.set(
            uuid,
            obs_path=str(obs_path.relative_to(config["OBSIDIAN_VAULT_DIR"])),
            title=extract_title(obs_path) or obs_path.stem,
            up_mtime=get_mtime(up_md_path),
            obs_mtime=obs_info["mtime"],
        )

    return True


# ─────────────────────────────────────────────
# 9. 메인 동기화 루프
# ─────────────────────────────────────────────
def run_sync(direction: str = "both", dry_run: bool = False):
    log.info("=" * 55)
    log.info(f"  UpNote ↔ Obsidian 동기화 시작  [{datetime.now():%Y-%m-%d %H:%M:%S}]")
    log.info(f"  모드: {direction}  |  dry_run: {dry_run}")
    log.info("=" * 55)

    sync_map = SyncMap(Path(CONFIG["SYNC_MAP_PATH"]))
    xl = ExcelLogger(Path(CONFIG["EXCEL_LOG_PATH"]))

    # 실행 시작 행 기록
    xl.record(
        mode=direction,
        status="DRY-RUN" if dry_run else "시작",
        title=f"총 스캔 시작",
    )

    # 스캔
    log.info("\n[1/4] UpNote 스캔 중...")
    upnote_notes = scan_upnote(CONFIG)
    log.info(f"  → {len(upnote_notes)}개 노트 발견")

    log.info("\n[2/4] Obsidian vault 스캔 중...")
    obsidian_notes = scan_obsidian(CONFIG)
    log.info(f"  → {len(obsidian_notes)}개 노트 발견 (UUID 추적)")

    # 모든 UUID 합집합
    all_uuids = set(upnote_notes) | set(obsidian_notes) | set(sync_map.all_uuids())

    log.info(f"\n[3/4] 동기화 처리 중... (총 {len(all_uuids)}개 노트)\n")

    stats = {"up_to_obs": 0, "obs_to_up": 0, "skip": 0, "error": 0}

    for uuid in sorted(all_uuids):
        up_info = upnote_notes.get(uuid)
        obs_info = obsidian_notes.get(uuid)

        action = decide_direction(uuid, up_info, obs_info, sync_map)

        # 충돌은 decide_direction에서 warning 로그 남기므로 엑셀에도 별도 기록
        if action in ("up_to_obs", "obs_to_up"):
            _prev = sync_map.get(uuid)
            _up_mt = up_info["mtime"] if up_info else 0.0
            _obs_mt = obs_info["mtime"] if obs_info else 0.0
            _is_conflict = (
                _prev is not None
                and _up_mt > _prev.get("up_mtime", 0.0) + CONFIG["MTIME_TOLERANCE_SEC"]
                and _obs_mt > _prev.get("obs_mtime", 0.0) + CONFIG["MTIME_TOLERANCE_SEC"]
            )
            if _is_conflict:
                xl.record(
                    mode=direction,
                    status="충돌",
                    uuid=uuid,
                    title=(up_info or obs_info or {}).get("title", ""),
                )

        try:
            if action in ("up_to_obs", "new_up") and direction in ("both", "up_to_obs"):
                if up_info:
                    ok = sync_up_to_obs(uuid, up_info, sync_map, CONFIG, dry_run)
                    stats["up_to_obs"] += int(ok)
                    xl.record(
                        mode=direction,
                        status="DRY-RUN" if dry_run else "UP→OBS",
                        uuid=uuid,
                        title=up_info.get("title", ""),
                    )
                else:
                    stats["skip"] += 1
                    xl.record(mode=direction, status="스킵", uuid=uuid)

            elif action in ("obs_to_up", "new_obs") and direction in ("both", "obs_to_up"):
                if obs_info:
                    ok = sync_obs_to_up(uuid, obs_info, up_info, sync_map, CONFIG, dry_run)
                    stats["obs_to_up"] += int(ok)
                    _title = (
                        up_info.get("title", "") if up_info
                        else extract_title(obs_info["obs_path"]) or obs_info["obs_path"].stem
                    )
                    xl.record(
                        mode=direction,
                        status="DRY-RUN" if dry_run else "OBS→UP",
                        uuid=uuid,
                        title=_title,
                    )
                else:
                    stats["skip"] += 1
                    xl.record(mode=direction, status="스킵", uuid=uuid)

            else:
                log.debug(f"  — skip [{uuid[:8]}]")
                stats["skip"] += 1
                # 스킵은 엑셀에 기록하지 않음 (행이 너무 많아짐)

        except Exception as e:
            log.error(f"  ❌ 오류 [{uuid[:8]}]: {e}")
            stats["error"] += 1
            xl.record(
                mode=direction,
                status="오류",
                uuid=uuid,
                title=(up_info or {}).get("title", ""),
                error=str(e),
            )

    # 맵 저장
    if not dry_run:
        sync_map.save()
        log.info(f"\n[4/4] sync_map.json 저장 완료")
    else:
        log.info(f"\n[4/4] [DRY-RUN] sync_map.json 저장 건너뜀")

    # 최종 요약 행 기록 후 엑셀 flush
    xl.record(
        mode=direction,
        status="완료",
        title=(
            f"UP→OBS {stats['up_to_obs']}건 | "
            f"OBS→UP {stats['obs_to_up']}건 | "
            f"스킵 {stats['skip']}건 | "
            f"오류 {stats['error']}건"
        ),
    )
    xl.flush()

    # 최종 요약
    log.info("\n" + "─" * 55)
    log.info(f"  ✅ UP→OBS: {stats['up_to_obs']}개  |  OBS→UP: {stats['obs_to_up']}개")
    log.info(f"  — 변경없음: {stats['skip']}개  |  ❌ 오류: {stats['error']}개")
    log.info("─" * 55 + "\n")


# ─────────────────────────────────────────────
# 10. CLI 진입점
# ─────────────────────────────────────────────
if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="UpNote → Obsidian 동기화")
    group = parser.add_mutually_exclusive_group()
    group.add_argument("--up-to-obs", action="store_true", help="UpNote → Obsidian 단방향 (기본)")
    group.add_argument(
        "--obs-to-up",
        action="store_true",
        help="[비활성화] UpNote가 외부 md를 자동 인식하지 않아 지원하지 않습니다.",
    )
    parser.add_argument("--dry-run", action="store_true", help="실제 변경 없이 미리보기")
    args = parser.parse_args()

    if args.obs_to_up:
        print(
            "\n[안내] --obs-to-up 은 현재 지원하지 않습니다.\n"
            "  UpNote 앱은 외부에서 생성된 md 파일을 자동으로 인식하지 않기 때문에,\n"
            "  Obsidian → UpNote 방향은 수동 Import 가 필요합니다.\n"
            "  현재는 UpNote → Obsidian 단방향만 운용합니다.\n"
        )
        raise SystemExit(0)

    direction = "up_to_obs"  # both 제거 — 항상 단방향
    run_sync(direction=direction, dry_run=args.dry_run)