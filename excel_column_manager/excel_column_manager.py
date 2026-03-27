"""
엑셀 컬럼 일괄 관리 스크립트
=============================
폴더 내 모든 .xlsx / .xls 파일에 대해
아래 [작업 목록]에 정의된 컬럼 작업을 순서대로 실행합니다.

지원 작업:
  - insert : 지정 위치에 새 컬럼 삽입
  - delete : 컬럼명으로 찾아서 삭제
  - rename : 컬럼명 변경

사전 설치:
    pip install openpyxl xlrd xlwt

사용법:
    python excel_column_manager.py
"""

import shutil
from pathlib import Path
from copy import copy

from openpyxl import load_workbook, Workbook

# ================================================================
# [기본 설정]
# ================================================================

# 처리할 폴더 경로
TARGET_FOLDER = r"."

# 원본 백업 폴더명 (None 이면 백업 안 함)
BACKUP_FOLDER = "backup_original"

# 헤더가 있는 행 번호 (1부터 시작)
HEADER_ROW = 1

# .xls → .xlsx 자동 변환 여부
CONVERT_XLS_TO_XLSX = True

# ================================================================
# [작업 목록]  ← 여기만 수정하면 됩니다
#
# 작업은 위에서 아래 순서대로 실행됩니다.
# 열 번호(col)는 작업 실행 시점의 실제 열 위치 기준입니다.
#
# ▸ 컬럼 삽입
#   {"op": "insert", "col": 23, "header": "I/F여부"}
#   → 23번째 열(W열)에 빈 컬럼 삽입, 헤더명 "I/F여부"
#
# ▸ 컬럼 삭제
#   {"op": "delete", "header": "삭제할컬럼명"}
#   → 해당 헤더명을 가진 컬럼을 찾아 삭제
#
# ▸ 컬럼명 변경
#   {"op": "rename", "header": "기존컬럼명", "new_header": "새컬럼명"}
#   → 해당 헤더명을 찾아 새 이름으로 변경
#
# ================================================================
COLUMN_TASKS = [
    {"op": "insert", "col": 23, "header": "I/F여부"},
    # {"op": "delete", "header": "불필요한컬럼"},
    # {"op": "rename", "header": "기존이름", "new_header": "새이름"},
]
# ================================================================


# ── 스타일 복사 ───────────────────────────────────────────────

def copy_cell_style(src, dst):
    if src.has_style:
        dst.font      = copy(src.font)
        dst.alignment = copy(src.alignment)
        dst.fill      = copy(src.fill)
        dst.border    = copy(src.border)
        dst.number_format = src.number_format


# ── 헤더 위치 탐색 ────────────────────────────────────────────

def find_col_by_header(sheet, header_name, header_row):
    """헤더명으로 열 번호(1-based)를 반환. 없으면 None."""
    for cell in sheet[header_row]:
        if cell.value == header_name:
            return cell.column
    return None


# ── 개별 작업 실행 ────────────────────────────────────────────

def op_insert(sheet, task, header_row):
    """지정 위치에 새 컬럼 삽입 후 헤더 입력."""
    col   = task["col"]
    hdr   = task["header"]

    sheet.insert_cols(col)

    new_cell = sheet.cell(row=header_row, column=col)
    new_cell.value = hdr

    # 왼쪽 셀 스타일 복사
    if col > 1:
        copy_cell_style(sheet.cell(row=header_row, column=col - 1), new_cell)

    return f"삽입 완료: {col}열 ← '{hdr}'"


def op_delete(sheet, task, header_row):
    """헤더명으로 컬럼을 찾아 삭제."""
    hdr = task["header"]
    col = find_col_by_header(sheet, hdr, header_row)

    if col is None:
        return f"삭제 건너뜀: '{hdr}' 컬럼을 찾을 수 없음"

    sheet.delete_cols(col)
    return f"삭제 완료: '{hdr}' (원래 {col}열)"


def op_rename(sheet, task, header_row):
    """헤더명을 찾아 새 이름으로 변경."""
    hdr     = task["header"]
    new_hdr = task["new_header"]
    col     = find_col_by_header(sheet, hdr, header_row)

    if col is None:
        return f"변경 건너뜀: '{hdr}' 컬럼을 찾을 수 없음"

    sheet.cell(row=header_row, column=col).value = new_hdr
    return f"변경 완료: '{hdr}' → '{new_hdr}' ({col}열)"


# 작업 유형 → 함수 매핑
OPERATION_MAP = {
    "insert": op_insert,
    "delete": op_delete,
    "rename": op_rename,
}


# ── 시트 단위 처리 ────────────────────────────────────────────

def apply_tasks_to_sheet(sheet, tasks, header_row):
    """작업 목록을 순서대로 시트에 적용."""
    for task in tasks:
        op   = task.get("op")
        func = OPERATION_MAP.get(op)
        if func is None:
            print(f"      [경고] 알 수 없는 작업 유형: '{op}' → 건너뜀")
            continue
        result = func(sheet, task, header_row)
        print(f"      {result}")


# ── 파일 단위 처리 ────────────────────────────────────────────

def process_xlsx(file_path: Path, backup_dir):
    if backup_dir:
        shutil.copy2(file_path, backup_dir / file_path.name)

    wb = load_workbook(file_path)
    for sheet_name in wb.sheetnames:
        print(f"    시트: [{sheet_name}]")
        apply_tasks_to_sheet(wb[sheet_name], COLUMN_TASKS, HEADER_ROW)

    wb.save(file_path)


def xls_to_openpyxl_workbook(xls_path: Path) -> Workbook:
    import xlrd
    xls_wb = xlrd.open_workbook(str(xls_path))
    new_wb = Workbook()
    new_wb.remove(new_wb.active)

    for sheet_name in xls_wb.sheet_names():
        xls_sheet = xls_wb.sheet_by_name(sheet_name)
        new_sheet = new_wb.create_sheet(title=sheet_name)
        for r in range(xls_sheet.nrows):
            for c in range(xls_sheet.ncols):
                new_sheet.cell(row=r + 1, column=c + 1, value=xls_sheet.cell(r, c).value)

    return new_wb


def process_xls(file_path: Path, backup_dir):
    if backup_dir:
        shutil.copy2(file_path, backup_dir / file_path.name)

    if CONVERT_XLS_TO_XLSX:
        wb = xls_to_openpyxl_workbook(file_path)
        for sheet_name in wb.sheetnames:
            print(f"    시트: [{sheet_name}]")
            apply_tasks_to_sheet(wb[sheet_name], COLUMN_TASKS, HEADER_ROW)

        save_path = file_path.with_suffix(".xlsx")
        wb.save(save_path)
        file_path.unlink()
        print(f"    → .xlsx 변환 저장: {save_path.name}")

    else:
        # xlwt 사용 (스타일 미지원, insert 작업만 지원)
        import xlrd, xlwt
        xls_wb = xlrd.open_workbook(str(file_path), formatting_info=True)
        new_wb = xlwt.Workbook()

        # insert 작업만 추출 (xlwt는 insert_cols 미지원 → 직접 재구성)
        insert_tasks = [t for t in COLUMN_TASKS if t["op"] == "insert"]
        if len(insert_tasks) != len(COLUMN_TASKS):
            print("    [경고] xls 유지 모드에서는 insert 작업만 지원됩니다.")

        for sheet_name in xls_wb.sheet_names():
            xls_sheet = xls_wb.sheet_by_name(sheet_name)
            new_sheet = new_wb.add_sheet(sheet_name)
            total_cols = xls_sheet.ncols
            insert_cols = sorted([t["col"] for t in insert_tasks])

            for r in range(xls_sheet.nrows):
                write_col = 0
                src_col   = 0
                for abs_col in range(total_cols + len(insert_cols)):
                    if (abs_col + 1) in insert_cols:
                        task  = next(t for t in insert_tasks if t["col"] == abs_col + 1)
                        value = task["header"] if r == HEADER_ROW - 1 else ""
                        new_sheet.write(r, write_col, value)
                    else:
                        if src_col < total_cols:
                            new_sheet.write(r, write_col, xls_sheet.cell(r, src_col).value)
                        src_col += 1
                    write_col += 1

        new_wb.save(str(file_path))


# ── 메인 ─────────────────────────────────────────────────────

def main():
    target = Path(TARGET_FOLDER).resolve()

    all_files = []
    for ext in ("*.xlsx", "*.xls"):
        for f in target.glob(ext):
            if BACKUP_FOLDER and BACKUP_FOLDER in f.parts:
                continue
            all_files.append(f)
    all_files = sorted(all_files)

    if not all_files:
        print(f"[알림] '{target}' 폴더에 처리할 엑셀 파일이 없습니다.")
        return

    backup_dir = None
    if BACKUP_FOLDER:
        backup_dir = target / BACKUP_FOLDER
        backup_dir.mkdir(exist_ok=True)
        print(f"[백업] 원본 파일 → '{backup_dir}'\n")

    # 실행할 작업 목록 출력
    print("[작업 목록]")
    for i, t in enumerate(COLUMN_TASKS, 1):
        op = t.get("op")
        if op == "insert":
            print(f"  {i}. 삽입  {t['col']}열 ← '{t['header']}'")
        elif op == "delete":
            print(f"  {i}. 삭제  '{t['header']}'")
        elif op == "rename":
            print(f"  {i}. 변경  '{t['header']}' → '{t['new_header']}'")
    print()

    xlsx_count = sum(1 for f in all_files if f.suffix.lower() == ".xlsx")
    xls_count  = sum(1 for f in all_files if f.suffix.lower() == ".xls")
    print(f"처리 대상: .xlsx {xlsx_count}개 / .xls {xls_count}개 (합계 {len(all_files)}개)\n")
    print("=" * 50)

    success, failed = 0, []

    for file_path in all_files:
        ext   = file_path.suffix.lower()
        label = "(xls→xlsx 변환)" if ext == ".xls" and CONVERT_XLS_TO_XLSX else ""
        print(f"처리 중: {file_path.name}  {label}")
        try:
            if ext == ".xlsx":
                process_xlsx(file_path, backup_dir)
            elif ext == ".xls":
                process_xls(file_path, backup_dir)
            print(f"  → 완료\n")
            success += 1
        except Exception as e:
            print(f"  → 오류 발생: {e}\n")
            failed.append((file_path.name, str(e)))

    print("=" * 50)
    print(f"\n[결과] 성공: {success}개 / 실패: {len(failed)}개")
    if failed:
        print("\n[실패 목록]")
        for name, err in failed:
            print(f"  - {name}: {err}")


if __name__ == "__main__":
    main()